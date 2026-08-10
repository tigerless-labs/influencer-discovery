from .gate import Gate
from .paths import memory_dir
from .record import Outcome

HEADERS = [
    "Name", "Email", "Email Source", "Followers", "Topic Evidence",
    "Verdict", "Profile", "Own Site", "Channel", "Added",
]
VERDICT = {
    Outcome.QUALIFIED: "qualified",
    Outcome.NO_CONTACT: "no contact",
    Outcome.BUYER: "buyer",
    Outcome.AUDIENCE_OUT_OF_BAND: "audience out of band",
    Outcome.AUDIENCE_UNVERIFIED: "audience unverified",
    Outcome.OFF_TOPIC: "off topic",
}
SUMMARY_TAB = "Summary"


def row_for(candidate):
    contact = next((c for c in candidate.contacts if c.kind == "email"), None)
    audience = candidate.audience
    return [
        candidate.display_name,
        contact.value if contact else "",
        contact.source if contact else "",
        audience.value if audience and audience.value else "",
        ", ".join(candidate.signals.get("topic_hits") or []),
        VERDICT.get(candidate.outcome, ""),
        candidate.profile_url or "",
        candidate.own_site or "",
        candidate.channel,
        (candidate.seen_at or "")[:10],
    ]


def contactable(store, band):
    """One row per person who can actually be written to. The verdict rides along as a column."""
    ranker = Gate(band)
    by_channel = {}
    for candidate in store.people():
        if not any(c.kind == "email" for c in candidate.contacts):
            continue
        if candidate.outcome is Outcome.AUDIENCE_OUT_OF_BAND:
            continue
        by_channel.setdefault(candidate.channel, []).append(candidate)
    return {
        channel: sorted(
            ranker.rank(people), key=lambda c: c.outcome is not Outcome.QUALIFIED
        )
        for channel, people in sorted(by_channel.items())
    }


def _style(sheet, widths):
    from openpyxl.styles import Alignment, Font, PatternFill

    header = sheet[1]
    for cell in header:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in zip(sheet.columns, widths):
        sheet.column_dimensions[column[0].column_letter].width = width


def to_workbook(by_channel):
    from openpyxl import Workbook

    widths = [26, 34, 15, 11, 26, 13, 34, 34, 12, 12]
    book = Workbook()
    summary = book.active
    summary.title = SUMMARY_TAB
    summary.append(["Channel", "With Email", "Qualified"])
    for channel, people in by_channel.items():
        summary.append([
            channel, len(people),
            sum(1 for c in people if c.outcome is Outcome.QUALIFIED),
        ])
    summary.append([
        "Total",
        sum(len(p) for p in by_channel.values()),
        sum(1 for p in by_channel.values() for c in p if c.outcome is Outcome.QUALIFIED),
    ])
    _style(summary, [16, 10, 11])

    for channel, people in by_channel.items():
        sheet = book.create_sheet(channel[:31])
        sheet.append(HEADERS)
        for candidate in people:
            sheet.append(row_for(candidate))
        _style(sheet, widths)
    return book


def write_xlsx(store, band, name="contactable"):
    """Real addresses, so it lands in the user's own directory and never in the repo."""
    directory = memory_dir()
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / f"{name}.xlsx"
    to_workbook(contactable(store, band)).save(path)
    return path
